# -*- coding: utf-8 -*-
"""학술/DB 신규 소스 정규화: CryptoScamDB / yuanqi / PTXPhish / SILENT-KILLER / Midsummer."""
import csv, io, json, os, re, zipfile
import yaml
import openpyxl
from collections import Counter

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
RAW = os.path.join(BASE, "raw")
OUT = os.path.join(BASE, "sources")

ADDR_BND = re.compile(r"(?<![0-9a-fA-F])0x[0-9a-fA-F]{40}(?![0-9a-fA-F])")
SOL_RE = re.compile(r"^[1-9A-HJ-NP-Za-km-z]{42,44}$")

def save(name, rows):
    with open(os.path.join(OUT, name), "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["address", "chain", "category", "source", "label", "detail", "ref_date"])
        w.writerows(rows)
    print(f"-> {name}: {len(rows)} rows  {dict(Counter(r[1] for r in rows))}")

# ---- CryptoScamDB (addresses: {chain:[...]}) ----
rows = []
data = yaml.safe_load(open(os.path.join(RAW, "cryptoscamdb_urls.yaml"), encoding="utf-8"))
for it in data:
    ad = it.get("addresses") or {}
    if not isinstance(ad, dict):
        continue
    cat = (it.get("category") or "").lower() or "scam"
    name = (it.get("name") or "")[:50]
    for ch, lst in ad.items():
        if ch not in ("ETH",):   # SOL/ETH만 관심 (여기선 ETH만 유의미)
            continue
        for a in (lst or []):
            if re.match(r"^0x[0-9a-fA-F]{40}$", a):
                rows.append([a.lower(), "ETH", "scam_" + cat, "cryptoscamdb", it.get("subcategory") or "", name, ""])
save("cryptoscamdb_eth.csv", rows)

# ---- yuanqi phishing-address.txt ----
rows = []
z = zipfile.ZipFile(os.path.join(RAW, "yuanqi_phishing.zip"))
txt = z.read("phishing accounts/phishing-address.txt").decode("utf-8", "replace")
for a in set(m.lower() for m in re.findall(r"0x[0-9a-fA-F]{40}", txt)):
    rows.append([a, "ETH", "phishing", "yuanqi_ethereum", "academic", "Network Embedding paper", ""])
save("yuanqi_eth.csv", rows)

# ---- PTXPhish (경계보정 주소 + InitialAddress seeds) ----
rows = set()
wb = openpyxl.load_workbook(os.path.join(RAW, "ptxphish.xlsx"))
for r in wb["Sheet1"].iter_rows(values_only=True):
    for v in r:
        if isinstance(v, str):
            for m in ADDR_BND.findall(v):
                rows.add(m.lower())
try:
    wb2 = openpyxl.load_workbook(os.path.join(RAW, "ptxphish_initial.xlsx"))
    for ws in wb2.worksheets:
        for r in ws.iter_rows(values_only=True):
            for v in r:
                if isinstance(v, str):
                    for m in ADDR_BND.findall(v):
                        rows.add(m.lower())
except Exception as e:
    print("  (InitialAddress 스킵:", e, ")")
save("ptxphish_eth.csv", [[a, "ETH", "phishing", "ptxphish_ndss2025", "academic", "BlockSec NDSS2025", ""] for a in sorted(rows)])

# ---- SILENT-KILLER (솔라나 rug dev + mint) ----
rows = []
z = zipfile.ZipFile(os.path.join(RAW, "silent_killer.zip"))
for line in z.open("rug_attacker_fingerprints.jsonl"):
    d = json.loads(line.decode())
    dev = d.get("dev_wallet", "")
    if dev and SOL_RE.match(dev):
        rows.append([dev, "SOL", "rugpull_dev", "silent_killer", f"rug_count={d.get('rug_count')}", "pump.fun rug dev", (d.get("last_rug_ts") or "")[:10]])
    for m in d.get("rugged_mints", []):
        mint = m.get("mint", "")
        if mint and SOL_RE.match(mint):
            rows.append([mint, "SOL", "rugpull_token", "silent_killer", m.get("symbol") or "", f"dev={dev}", (m.get("gone_ts") or "")[:10]])
# 중복 mint 제거
seen = set(); dedup = []
for r in rows:
    if r[0] in seen: continue
    seen.add(r[0]); dedup.append(r)
save("silent_killer_sol.csv", dedup)

# ---- Midsummer Meme (솔라나 wash-trade maker 지갑 — 저신뢰) ----
zm = zipfile.ZipFile(os.path.join(RAW, "midsummer_meme.zip"))
addrs = set()
with zm.open("meme_coin_anon-main/data/dune_data_on_potential_wash_trading_makers_HP_coins.csv") as fh:
    for r in csv.DictReader(io.TextIOWrapper(fh, encoding="utf-8")):
        if (r.get("platform") or "").lower() != "solana":
            continue
        for col in ("maker", "address"):
            a = (r.get(col) or "").strip()
            if a and SOL_RE.match(a):
                addrs.add(a)
mm = [[a, "SOL", "wash_trading_suspect", "midsummer_meme", "heuristic", "Zenodo 17830944", ""] for a in sorted(addrs)]
save("midsummer_wash_sol.csv", mm)
